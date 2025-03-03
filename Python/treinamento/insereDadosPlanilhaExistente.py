import openpyxl

def criar_excel(nome_arquivo):
    """Cria um novo arquivo Excel e adiciona dados."""
    # Criar um novo workbook e uma planilha ativa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Adicionando cabeçalhos
    ws.append(["ID", "Nome", "Idade"])

    # Adicionando alguns dados de exemplo
    dados = [
        [1, "Anderson", 30],
        [2, "Maria", 25],
        [3, "Carlos", 40]
    ]

    for linha in dados:
        ws.append(linha)
    
    # Escrever na coluna M, posição 3
    ws["M3"] = "Novo Dado"

    # Salvando o arquivo
    wb.save(nome_arquivo)
    print(f"Arquivo Excel '{nome_arquivo}' criado com sucesso!")

def ler_dados_excel(nome_arquivo):
    """Lê dados de um arquivo Excel."""
    wb = openpyxl.load_workbook(nome_arquivo)
    ws = wb.active
    
    dados = []
    for row in ws.iter_rows(values_only=True):
        dados.append(row)
    
    print(" Dados do Excel:")
    for linha in dados:
        print(linha)
    
    return dados

def escrever_celula(nome_arquivo, celula, valor):
    """Escreve um valor em uma célula específica da planilha existente."""
    wb = openpyxl.load_workbook(nome_arquivo)
    ws = wb.active
    ws[celula] = valor
    wb.save(nome_arquivo)
    print(f" Valor '{valor}' escrito na célula {celula} do arquivo '{nome_arquivo}'.")

if __name__ == "__main__":
    nome_arquivo = "conexaoTerceirosComAE.xlsx"
    #criar_excel(nome_arquivo)
    #ler_dados_excel(nome_arquivo)
    escrever_celula(nome_arquivo, "A9", "excell")
